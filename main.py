from pathlib import Path
import datetime
import re
import icalendar
import openpyxl
import zoneinfo

PATH_FOLDER = Path(__file__).parent
PATH_RAW_CAL = PATH_FOLDER / 'EDT_raw.xlsx'
PATH_CAL_EXPORT = PATH_FOLDER / 'EDT.ics'
TIMEZONE = zoneinfo.ZoneInfo('Europe/Paris')

RE_TIME = re.compile(r'(?P<hours1>\d{1,2})[hH]\s?(?P<minutes1>\d{1,2})?\s?[-/]\s?(?P<hours2>\d{1,2})[hH]\s?(?P<minutes2>\d{1,2})?')

class Event:
    def __init__(self, date: datetime.date, start_time: str, end_time: str, description: str):
        self.date = date
        self.start_time = start_time
        self.end_time = end_time
        self.description = description

    def __repr__(self):
        return f"Event(date={self.date}, start_time={self.start_time}, end_time={self.end_time}, description={self.description.replace('\n', '\t')})"

    def make_ical_event(self) -> icalendar.Event:
        event = icalendar.Event()
        start_dt = datetime.datetime.combine(self.date, datetime.datetime.strptime(self.start_time, "%H:%M").time(), TIMEZONE)
        end_dt = datetime.datetime.combine(self.date, datetime.datetime.strptime(self.end_time, "%H:%M").time(), TIMEZONE)
        event.add('dtstart', start_dt)
        event.add('dtend', end_dt)
        event.add('summary', self.description)
        return event

def main():
    # Import and read the Excel file
    wb = openpyxl.load_workbook(PATH_RAW_CAL)
    sheet = wb.active
    if not sheet:
        raise RuntimeError("No active sheet found in the Excel file.")

    # Start rows for each weeks
    week_rows = [3 + 9*i for i in range(14 + 1)] + \
                [155 + 9*i for i in range(11 + 1)]
     # [138, 146] + \

    # Iterate over each week
    events:list[Event] = []
    for row_start in week_rows:
        # === Weekly parsing ===
        raw_monday_date = sheet.cell(row=row_start, column=1).value
        # print(f"Raw monday date type: {type(raw_monday_date)}, value: {raw_monday_date}")
        
        # Convert to date object
        if isinstance(raw_monday_date, datetime.datetime):
            monday_date = raw_monday_date.date()
        elif isinstance(raw_monday_date, datetime.date):
            monday_date = raw_monday_date
        elif isinstance(raw_monday_date, str):
            # Handle string format if needed
            monday_date = datetime.datetime.strptime(raw_monday_date, "%Y-%m-%d").date()
        else:
            raise ValueError(f"Unexpected date format: {type(raw_monday_date)} - {raw_monday_date}")
        # Force 2025
        monday_date = monday_date.replace(year=2025)
            
        # print(f"Week starting on: {monday_date}")
        for col in range(2, 11 + 1):
            for row in (row_start, row_start+4): # Morning / Afternoon
                # print(f"{col}: {sheet.cell(row=row, column=col).value}")
                day_offset = (col-1)//2
                day = monday_date + datetime.timedelta(days=day_offset)

                contents = [sheet.cell(row=row+i, column=col).value for i in range(4)]
                content = "\n".join([str(c) for c in contents if c])
                if not content:
                    continue
                # if content:
                #     print(f"  {day} ({'AM' if row==row_start else 'PM'}) : {str(content).replace('\n','\t')}")
                # else:
                #     print(f"  {day} ({'AM' if row==row_start else 'PM'})")

                # == Create event ==
                m = RE_TIME.search(content)
                if m:
                    time_range = (m.group('hours1'), m.group('minutes1') or '00', m.group('hours2'), m.group('minutes2') or '00')
                    # print(f"  Found time range: {time_range[0]}:{time_range[1]} to {time_range[2]}:{time_range[3]}")
                else:
                    if row == row_start: # morning
                        time_range = ('07', '00', '13', '00') 
                    else:
                        time_range = ('13', '00', '19', '00')
                    content += "(HEURES NON SPECIFIEES)"
                    # print(f"  WARNING: No time range found for entry: {content}, assuming {time_range[0]}:{time_range[1]} to {time_range[2]}:{time_range[3]}")
                        
                start_time = f"{time_range[0]}:{time_range[1]}"
                end_time = f"{time_range[2]}:{time_range[3]}"
                event = Event(date=day, start_time=start_time, end_time=end_time, description=content)
                events.append(event)
                
    # Export all events
    cal = icalendar.Calendar()
    cal.add('prodid', '-//SETIiCal//SETIiCal//EN')
    cal.add('version', '2.0')
    for e in events:
        cal.add_component(e.make_ical_event())

    with open(PATH_CAL_EXPORT, 'wb') as f:
        f.write(cal.to_ical())

    # Export events by type
    events_by_type: dict[str, list[Event]] = {}
    for e in events:
        if (e.description.startswith("A0") or 
        e.description.startswith("B0") or 
        e.description.startswith("C0") or
        e.description.startswith("RaN")):
            typ = "TC"
        elif e.description.startswith("A1"):
            typ = "A1"
        elif e.description.startswith("A2"):
            typ = "A2"
        elif e.description.startswith("A3"):
            typ = "A3"
        elif e.description.startswith("A4"):
            typ = "A4"
        elif e.description.startswith("A5"):
            typ = "A5"
        elif e.description.startswith("B1"):
            typ = "B1"
        elif e.description.startswith("B2"):
            typ = "B2"
        elif e.description.startswith("B3"):
            typ = "B3"
        elif e.description.startswith("B4"):
            typ = "B4"
        elif e.description.startswith("B5"):
            typ = "B5"
        elif e.description.startswith("C1"):
            typ = "C1"
        elif e.description.startswith("C2"):
            typ = "C2"
        elif e.description.startswith("C3"):
            typ = "C3"
        elif e.description.startswith("C4"):
            typ = "C4"
        elif e.description.startswith("C5"):
            typ = "C5"
        elif e.description.startswith("IDG"):
            typ = "IDG"
        elif e.description.startswith("IR"):
            typ = "IR"
        else:
            typ = "OTHER"
            print(f"WARNING: no type found for: {e=}")

        if typ not in events_by_type:
            events_by_type[typ] = []
        events_by_type[typ].append(e)

    for t in events_by_type:
        cal = icalendar.Calendar()
        cal.add('prodid', '-//SETIiCal//SETIiCal//EN')
        cal.add('version', '2.0')
        for e in events_by_type[t]:
            cal.add_component(e.make_ical_event())
        path_export = PATH_FOLDER / f"EDT_{t}.ics"
        with open(path_export, 'wb') as f:
            f.write(cal.to_ical())

if __name__ == '__main__':
    main()